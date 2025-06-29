import os
import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict, Any


class DataProcessor:
    """数据处理模块，负责Excel文件的读写操作"""
    
    # Excel列定义
    EXCEL_COLUMNS_MATCHED = [
        "bgm_id", "bgm游戏", "日文名 (原始)", "中文名 (原始)",
        "name", "chineseName", "ym_id", "score",
        "orgId", "orgName", "orgWebsite", "orgDescription",
        "匹配来源"
    ]
    
    EXCEL_COLUMNS_ORG = [
        "org_id", "name", "chineseName", "website", "description", "birthday"
    ]
    
    def __init__(self):
        pass
    
    def init_excel(self, output_file: str) -> None:
        """
        确保匹配结果文件存在；若不存在或损坏则创建带表头的新文件。
        """
        need_create = False
        if not os.path.exists(output_file):
            need_create = True
        else:
            try:
                load_workbook(output_file)
            except Exception:
                need_create = True

        if need_create:
            pd.DataFrame(columns=self.EXCEL_COLUMNS_MATCHED).to_excel(output_file, index=False)
            print(f"已初始化输出文件：{output_file}")

    def init_org_excel(self, output_file: str) -> None:
        """类似 ``init_excel``，但针对会社信息文件。"""
        if not os.path.exists(output_file):
            pd.DataFrame(columns=self.EXCEL_COLUMNS_ORG).to_excel(output_file, index=False)
            print(f"已初始化会社信息文件：{output_file}")

    def append_to_excel(self, row_data: List[Dict[str, Any]], output_file: str) -> None:
        """
        将 ``row_data`` 追加写入到 ``output_file``，支持自动创建及占用兜底。
        """
        try:
            df_new = pd.DataFrame(row_data)

            # 1️⃣ 文件不存在：直接写
            if not os.path.exists(output_file):
                df_new.to_excel(output_file, index=False)
                return

            # 2️⃣ 文件存在：读取 + 合并 + 写回
            try:
                df_exist = pd.read_excel(output_file)
                df_combined = pd.concat([df_exist, df_new], ignore_index=True)
                df_combined.to_excel(output_file, index=False)
            except PermissionError:  # 常见于文件被 Excel 占用
                temp_file = f"{output_file}.temp"
                df_new.to_excel(temp_file, index=False)
                print(f"原文件被占用，数据已保存到临时文件：{temp_file}")
        except Exception as exc:
            # 兜底打印 & 备份
            print(f"保存数据时发生错误: {exc}")
            backup_file = f"{output_file}.backup"
            pd.DataFrame(row_data).to_excel(backup_file, index=False)
            print(f"数据已保存到备用文件：{backup_file}")

    def append_unmatched_to_excel(self, name: str, unmatched_file: str) -> None:
        """记录未匹配成功的 Bangumi 名称。"""
        df = pd.DataFrame([[name]], columns=["原始的未匹配bgm游戏名称"])
        if not os.path.exists(unmatched_file):
            df.to_excel(unmatched_file, index=False)
        else:
            with pd.ExcelWriter(unmatched_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                sheet = writer.book["Sheet1"]
                df.to_excel(writer, index=False, header=False, startrow=sheet.max_row)

    def append_org_to_excel(self, org_info: Dict[str, Any], output_file: str) -> None:
        """将会社信息写入文件，逻辑同 ``append_to_excel``。"""
        self.append_to_excel([org_info], output_file)
    
    def read_bgm_data(self, input_file: str) -> pd.DataFrame:
        """读取Bangumi源文件"""
        df_bgm = pd.read_excel(input_file, engine="openpyxl")
        print(f"DEBUG: 识别到的 Excel 列名：{df_bgm.columns.tolist()}")
        
        if "日文名" not in df_bgm.columns or "中文名" not in df_bgm.columns:
            raise ValueError("Excel 中必须包含 '日文名' 和 '中文名' 列")
        
        return df_bgm
    
    def read_bgm_data_with_aliases(self, input_file: str) -> pd.DataFrame:
        """读取包含别名的Bangumi源文件"""
        df_bgm = pd.read_excel(input_file, engine="openpyxl")
        print(f"DEBUG: 识别到的 Excel 列名：{df_bgm.columns.tolist()}")
        return df_bgm
    
    def get_processed_ids(self, output_file: str) -> set:
        """获取已处理的ID集合，用于断点续跑"""
        processed_ids = set()
        if os.path.exists(output_file):
            try:
                df_exist = pd.read_excel(output_file, engine="openpyxl")
                if 'bgm_id' in df_exist.columns:
                    processed_ids = set(df_exist["bgm_id"].dropna().astype(str))
                else:
                    print("警告: 输出文件中未找到 'bgm_id' 列，断点续跑可能不准确。")
            except Exception as exc:
                print("读取已匹配文件失败，将重新创建：", exc)
        return processed_ids
    
    def get_processed_orgs(self, org_output_file: str) -> Dict[str, Dict[str, Any]]:
        """获取已处理的会社信息，用于避免重复查询"""
        processed_orgs = {}
        if os.path.exists(org_output_file):
            try:
                org_df = pd.read_excel(org_output_file, engine="openpyxl")
                for _, row in org_df.iterrows():
                    org_id = str(row["org_id"])
                    if pd.notna(org_id):
                        processed_orgs[org_id] = {
                            "info": row.to_dict(),
                            "retry_count": 0
                        }
            except Exception as exc:
                print("读取会社信息文件失败，将重新创建：", exc)
        return processed_orgs 